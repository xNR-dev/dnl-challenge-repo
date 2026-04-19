#!/usr/bin/env python3
"""Generate an Excel checklist from checklist.json."""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT = os.path.join(SCRIPT_DIR, "checklist.json")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
OUTPUT = os.path.join(OUTPUT_DIR, "checklist.xlsx")

SECTION_FILLS = {
    "Bilanz": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "GuV": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Anhang": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "Lagebericht": PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
}

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_COLOR = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def main():
    with open(INPUT, "r", encoding="utf-8") as f:
        data = json.load(f)

    items = data["checklist"]
    meta = data.get("metadata", {})
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()

    # --- Overview sheet ---
    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_overview.append(["HGB Disclosure Checklist"])
    ws_overview.append([])
    
    meta_fields = [
        "entity_type",
        "size_class",
        "reporting_framework",
        "pl_format",
        "target_fy",
        "target_fy_effective_date",
        "framework_version",
        "schema_version",
    ]
    for key in meta_fields:
        if key in meta:
            ws_overview.append([key.replace("_", " ").title(), meta.get(key, "")])
    
    ws_overview.append([])
    ws_overview.append(["Total Items", len(items)])
    ws_overview.append(["Mandatory", sum(1 for i in items if i.get("obligation") == "M")])
    ws_overview.append(["Conditional", sum(1 for i in items if i.get("obligation") == "C")])

    # --- Per-section sheets ---
    COLUMNS = [
        "ID", "Section", "Sub-section", "Checklist Item",
        "HGB Reference", "Effective Version Date", "Historical Source URL",
        "Official Reference URL", "Source Text (DE)", "Obligation", "Trigger Condition",
        "Verification Mode", "Presence Inference Rule", "Cross Evidence Sources",
        "Evidence Location", "Completeness Prompt", "Audit Response"
    ]

    sections_order = ["Bilanz", "GuV", "Anhang", "Lagebericht"]
    for sec in sections_order:
        ws = wb.create_sheet(title=sec)

        # Header row
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT_COLOR
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Data rows
        sec_items = [i for i in items if i.get("section") == sec]
        for row_idx, item in enumerate(sec_items, 2):
            fill = SECTION_FILLS.get(sec, PatternFill())
            values = [
                item.get("id", ""),
                item.get("section", ""),
                item.get("sub_section", ""),
                item.get("disclosure_item", ""),
                item.get("hgb_reference", ""),
                item.get("effective_version_date", ""),
                item.get("historical_source_url", ""),
                item.get("official_reference_url", ""),
                item.get("source_text_de", ""),
                item.get("obligation", ""),
                item.get("trigger_condition", "") or "",
                item.get("verification_mode", ""),
                item.get("presence_inference_rule", ""),
                ", ".join(item.get("cross_evidence_sources", [])),
                item.get("evidence_location", ""),
                item.get("completeness_prompt", ""),
                "",  # Audit Response — blank for auditor
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col_idx == 1:
                    cell.fill = fill

        # Column widths
        widths = [12, 10, 20, 48, 22, 18, 32, 32, 36, 12, 28, 20, 36, 26, 52, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else None].width = w

    # --- Combined sheet ---
    ws_all = wb.create_sheet(title="All Items")
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws_all.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT_COLOR
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, item in enumerate(items, 2):
        values = [
            item.get("id", ""),
            item.get("section", ""),
            item.get("sub_section", ""),
            item.get("disclosure_item", ""),
            item.get("hgb_reference", ""),
            item.get("effective_version_date", ""),
            item.get("historical_source_url", ""),
            item.get("official_reference_url", ""),
            item.get("source_text_de", ""),
            item.get("obligation", ""),
            item.get("trigger_condition", "") or "",
            item.get("verification_mode", ""),
            item.get("presence_inference_rule", ""),
            ", ".join(item.get("cross_evidence_sources", [])),
            item.get("evidence_location", ""),
            item.get("completeness_prompt", ""),
            "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [12, 10, 20, 48, 22, 18, 32, 32, 36, 12, 28, 20, 36, 26, 52, 14]
    for i, w in enumerate(widths, 1):
        ws_all.column_dimensions[chr(64 + i) if i <= 26 else None].width = w

    wb.save(OUTPUT)
    print(f"Excel saved to {OUTPUT}")
    print(f"Total items: {len(items)}")
    for sec in sections_order:
        count = sum(1 for i in items if i.get("section") == sec)
        print(f"  {sec}: {count}")


if __name__ == "__main__":
    main()
