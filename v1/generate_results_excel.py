#!/usr/bin/env python3
"""Generate results Excel from checklist.json + results.json."""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CHECKLIST = os.path.join(SCRIPT_DIR, "checklist.json")
RESULTS = os.path.join(SCRIPT_DIR, "results.json")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
OUTPUT = os.path.join(OUTPUT_DIR, "results.xlsx")

RESPONSE_FILLS = {
    "Yes": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "No": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "N/A": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "Review": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

COLUMNS = [
    "ID", "Sub-section", "Checklist Item", "HGB Ref",
    "Oblig", "Trigger Condition", "Response", "Evidence", "Notes"
]
WIDTHS = [10, 20, 55, 22, 8, 30, 10, 50, 30]


def main():
    with open(CHECKLIST, "r", encoding="utf-8") as f:
        checklist = json.load(f)
    with open(RESULTS, "r", encoding="utf-8") as f:
        results = json.load(f)

    # Build results lookup
    res_map = {r["id"]: r for r in results}
    items = checklist["checklist"]
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()

    # --- Summary sheet ---
    ws_sum = wb.active
    ws_sum.title = "Summary"

    stats = {"Yes": 0, "No": 0, "N/A": 0, "Review": 0}
    sec_stats = {}
    issues = []

    for item in items:
        iid = item["id"]
        sec = item["section"]
        resp = res_map.get(iid, {}).get("response", "Review")
        stats[resp] = stats.get(resp, 0) + 1

        if sec not in sec_stats:
            sec_stats[sec] = {"Yes": 0, "No": 0, "N/A": 0, "Review": 0}
        sec_stats[sec][resp] = sec_stats[sec].get(resp, 0) + 1

        if resp in ("No", "Review"):
            issues.append({
                "id": iid, "section": sec,
                "question": item["disclosure_item"],
                "response": resp,
                "evidence": res_map.get(iid, {}).get("evidence", ""),
                "notes": res_map.get(iid, {}).get("notes", ""),
            })

    total = len(items)
    applicable = stats["Yes"] + stats["No"]
    pass_rate = (stats["Yes"] / applicable * 100) if applicable > 0 else 0

    ws_sum.append(["HGB Disclosure Checklist — Verification Results"])
    ws_sum.append(["Entity", checklist["metadata"].get("entity_type", "")])
    ws_sum.append(["Reporting Date", checklist["metadata"].get("applicable_date", "")])
    ws_sum.append([])
    ws_sum.append(["OVERALL RESULTS"])
    ws_sum.append(["Total Items", total])
    ws_sum.append(["Yes (disclosed)", stats["Yes"]])
    ws_sum.append(["No (missing)", stats["No"]])
    ws_sum.append(["N/A (not applicable)", stats["N/A"]])
    ws_sum.append(["Review (uncertain)", stats["Review"]])
    ws_sum.append(["Pass Rate (Yes / applicable)", f"{pass_rate:.1f}%"])
    ws_sum.append([])
    ws_sum.append(["BY SECTION"])
    ws_sum.append(["Section", "Total", "Yes", "No", "N/A", "Review", "Pass Rate"])

    sections_order = ["Bilanz", "GuV", "Anhang", "Lagebericht"]
    for sec in sections_order:
        s = sec_stats.get(sec, {"Yes": 0, "No": 0, "N/A": 0, "Review": 0})
        sec_total = sum(s.values())
        sec_applicable = s["Yes"] + s["No"]
        sec_rate = (s["Yes"] / sec_applicable * 100) if sec_applicable > 0 else 0
        ws_sum.append([sec, sec_total, s["Yes"], s["No"], s["N/A"], s["Review"], f"{sec_rate:.1f}%"])

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15

    # --- Section sheets ---
    for sec in sections_order:
        ws = wb.create_sheet(title=sec)

        # Header
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        sec_items = [i for i in items if i["section"] == sec]
        for row_idx, item in enumerate(sec_items, 2):
            res = res_map.get(item["id"], {})
            values = [
                item["id"],
                item.get("sub_section", ""),
                item["disclosure_item"],
                item["hgb_reference"],
                item["obligation"],
                item.get("trigger_condition", "") or "",
                res.get("response", "Review"),
                res.get("evidence", ""),
                res.get("notes", ""),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Colour code Response column
                if col_idx == 7:
                    cell.fill = RESPONSE_FILLS.get(val, PatternFill())

        for i, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[chr(64 + i) if i <= 9 else None].width = w
        # Fix column widths for columns beyond I
        col_letters = ["A","B","C","D","E","F","G","H","I"]
        for i, w in enumerate(WIDTHS):
            ws.column_dimensions[col_letters[i]].width = w

    # --- Issues sheet ---
    ws_issues = wb.create_sheet(title="Issues - No Review")
    issue_cols = ["ID", "Section", "Checklist Item", "Response", "Evidence", "Notes"]
    for col_idx, col_name in enumerate(issue_cols, 1):
        cell = ws_issues.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, issue in enumerate(issues, 2):
        values = [issue["id"], issue["section"], issue["question"],
                  issue["response"], issue["evidence"], issue["notes"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws_issues.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 4:
                cell.fill = RESPONSE_FILLS.get(val, PatternFill())

    issue_widths = [10, 12, 55, 10, 50, 30]
    issue_letters = ["A","B","C","D","E","F"]
    for i, w in enumerate(issue_widths):
        ws_issues.column_dimensions[issue_letters[i]].width = w

    wb.save(OUTPUT)
    print(f"Results Excel saved to {OUTPUT}")
    print(f"Total: {total} | Yes: {stats['Yes']} | No: {stats['No']} | N/A: {stats['N/A']} | Review: {stats['Review']}")
    print(f"Pass rate: {pass_rate:.1f}%")
    if issues:
        print(f"\nIssues ({len(issues)}):")
        for iss in issues:
            print(f"  {iss['response']:6} | {iss['id']} | {iss['question'][:60]}")


if __name__ == "__main__":
    main()
