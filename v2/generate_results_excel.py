#!/usr/bin/env python3
"""Generate results Excel from checklist.json + results.json."""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CHECKLIST = os.path.join(SCRIPT_DIR, "checklist.json")
RESULTS = os.path.join(SCRIPT_DIR, "results.json")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
OUTPUT = os.path.join(OUTPUT_DIR, "results.xlsx")

RESPONSE_FILLS = {
    "Yes": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "No": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "N/A": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "Review": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

REVIEW_COLUMNS = [
    "ID", "Sub-section", "Checklist Item", "HGB Ref", "Source Text (DE)",
    "Obligation", "Trigger Condition", "Evidence Location", "Response",
    "Reason Code", "Confidence", "Evidence", "Audit Notes"
]
REVIEW_WIDTHS = [12, 22, 42, 18, 34, 10, 24, 18, 12, 18, 12, 42, 30]

TRACEABILITY_COLUMNS = [
    "ID", "Section", "Sub-section", "Checklist Item", "HGB Ref",
    "Effective Version Date", "Historical Source URL", "Official Reference URL",
    "Verification Mode", "Presence Inference Rule", "Cross Evidence Sources",
    "Verified By", "Verification Date"
]
TRACEABILITY_WIDTHS = [12, 10, 22, 42, 18, 18, 28, 28, 18, 34, 24, 18, 15]


def main():
    with open(CHECKLIST, "r", encoding="utf-8") as f:
        checklist = json.load(f)
    with open(RESULTS, "r", encoding="utf-8") as f:
        results = json.load(f)

    # Build results lookup by ID
    res_map = {r["id"]: r for r in results}
    items = checklist["checklist"]
    meta = checklist.get("metadata", {})
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()

    # --- Summary sheet ---
    ws_sum = wb.active
    ws_sum.title = "Summary"

    stats = {"Yes": 0, "No": 0, "N/A": 0, "Review": 0}
    sec_stats = {}
    issues = []

    for item in items:
        iid = item["id"]
        sec = item.get("section", "Unknown")
        resp = res_map.get(iid, {}).get("response", "Review")
        stats[resp] = stats.get(resp, 0) + 1

        if sec not in sec_stats:
            sec_stats[sec] = {"Yes": 0, "No": 0, "N/A": 0, "Review": 0}
        sec_stats[sec][resp] = sec_stats[sec].get(resp, 0) + 1

        if resp in ("No", "Review"):
            issues.append({
                "id": iid,
                "section": sec,
                "question": item.get("disclosure_item", ""),
                "response": resp,
                "evidence": res_map.get(iid, {}).get("evidence", ""),
                "notes": res_map.get(iid, {}).get("audit_notes", ""),
            })

    total = len(items)
    applicable = stats["Yes"] + stats["No"]
    pass_rate = (stats["Yes"] / applicable * 100) if applicable > 0 else 0

    ws_sum.append(["HGB Disclosure Checklist — Verification Results"])
    ws_sum.append(["Entity", meta.get("entity_type", "")])
    ws_sum.append(["Reporting Date", meta.get("target_fy", meta.get("applicable_date", ""))])
    ws_sum.append(["Framework Version", meta.get("framework_version", "")])
    ws_sum.append(["Schema Version", meta.get("schema_version", "")])
    ws_sum.append([])
    ws_sum.append(["OVERALL RESULTS"])
    ws_sum.append(["Total Items", total])
    ws_sum.append(["Yes (disclosed)", stats["Yes"]])
    ws_sum.append(["No (missing)", stats["No"]])
    ws_sum.append(["N/A (not applicable)", stats["N/A"]])
    ws_sum.append(["Review (uncertain)", stats["Review"]])
    ws_sum.append(["Pass Rate (Yes / applicable)", f"{pass_rate:.1f}%"])
    ws_sum.append([])
    ws_sum.append(["BY SECTION"])
    ws_sum.append(["Section", "Total", "Yes", "No", "N/A", "Review", "Pass Rate"])

    sections_order = ["Bilanz", "GuV", "Anhang", "Lagebericht"]
    for sec in sections_order:
        s = sec_stats.get(sec, {"Yes": 0, "No": 0, "N/A": 0, "Review": 0})
        sec_total = sum(s.values())
        sec_applicable = s["Yes"] + s["No"]
        sec_rate = (s["Yes"] / sec_applicable * 100) if sec_applicable > 0 else 0
        ws_sum.append([sec, sec_total, s["Yes"], s["No"], s["N/A"], s["Review"], f"{sec_rate:.1f}%"])

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15

    # --- Reviewer-facing section sheets ---
    for sec in sections_order:
        ws = wb.create_sheet(title=sec)

        # Header
        for col_idx, col_name in enumerate(REVIEW_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        sec_items = [i for i in items if i.get("section") == sec]
        for row_idx, item in enumerate(sec_items, 2):
            res = res_map.get(item["id"], {})
            values = [
                item.get("id", ""),
                item.get("sub_section", ""),
                item.get("disclosure_item", ""),
                item.get("hgb_reference", ""),
                item.get("source_text_de", ""),
                item.get("obligation", ""),
                item.get("trigger_condition", "") or "",
                item.get("evidence_location", ""),
                res.get("response", "Review"),
                res.get("reason_code", ""),
                res.get("confidence", ""),
                res.get("evidence", ""),
                res.get("audit_notes", ""),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Colour code Response column
                if REVIEW_COLUMNS[col_idx - 1] == "Response":
                    cell.fill = RESPONSE_FILLS.get(val, PatternFill())

        for i, w in enumerate(REVIEW_WIDTHS):
            col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
            ws.column_dimensions[col_letter].width = w

    # --- Traceability sheet ---
    ws_trace = wb.create_sheet(title="Traceability")
    for col_idx, col_name in enumerate(TRACEABILITY_COLUMNS, 1):
        cell = ws_trace.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, item in enumerate(items, 2):
        res = res_map.get(item["id"], {})
        values = [
            item.get("id", ""),
            item.get("section", ""),
            item.get("sub_section", ""),
            item.get("disclosure_item", ""),
            item.get("hgb_reference", ""),
            item.get("effective_version_date", ""),
            item.get("historical_source_url", ""),
            item.get("official_reference_url", ""),
            item.get("verification_mode", ""),
            item.get("presence_inference_rule", ""),
            ", ".join(item.get("cross_evidence_sources", [])),
            res.get("verified_by", ""),
            res.get("verification_date", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_trace.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i, w in enumerate(TRACEABILITY_WIDTHS):
        col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
        ws_trace.column_dimensions[col_letter].width = w

    # --- Issues sheet ---
    ws_issues = wb.create_sheet(title="Issues - No Review")
    issue_cols = ["ID", "Section", "Checklist Item", "Response", "Reason Code", "Evidence", "Audit Notes"]
    issue_widths = [12, 10, 50, 10, 18, 50, 35]
    
    for col_idx, col_name in enumerate(issue_cols, 1):
        cell = ws_issues.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, issue in enumerate(issues, 2):
        res = res_map.get(issue["id"], {})
        values = [
            issue["id"],
            issue["section"],
            issue["question"],
            issue["response"],
            res.get("reason_code", ""),
            issue["evidence"],
            issue["notes"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_issues.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 4:
                cell.fill = RESPONSE_FILLS.get(val, PatternFill())

    for i, w in enumerate(issue_widths):
        ws_issues.column_dimensions[chr(65 + i)].width = w

    wb.save(OUTPUT)
    print(f"Results Excel saved to {OUTPUT}")
    print(f"Total: {total} | Yes: {stats['Yes']} | No: {stats['No']} | N/A: {stats['N/A']} | Review: {stats['Review']}")
    print(f"Pass rate: {pass_rate:.1f}%")
    if issues:
        print(f"\nIssues ({len(issues)}):")
        for iss in issues:
            print(f"  {iss['response']:6} | {iss['id']} | {iss['question'][:60]}")


if __name__ == "__main__":
    main()
